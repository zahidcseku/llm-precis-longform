import openpyxl
from typing import List, Any, Optional
import pandas as pd


def read_first_sheet_with_openpyxl(
    file_path: str,
) -> Optional[pd.DataFrame]:  # Optional[List[List[Any]]]:
    """
    Reads all data from the first sheet of an Excel file using the openpyxl library.

    Args:
        file_path: The path to the .xlsx file.

    Returns:
        A list of lists, where each inner list represents a row
        and contains the cell values of that row.
        Returns None if the file is not found, the sheet cannot be accessed,
        or another error occurs.
    """
    try:
        # Load the workbook. read_only=True is more efficient if you're only reading.
        workbook = openpyxl.load_workbook(filename=file_path, read_only=True)

        # Access the first sheet.
        # Option 1: Get the active sheet (usually the first sheet when a file is opened).
        first_sheet = workbook.active

        # Option 2: Explicitly get the first sheet by its index in the sheetnames list.
        # if workbook.sheetnames:
        #     first_sheet_name = workbook.sheetnames[0]
        #     first_sheet = workbook[first_sheet_name]
        # else:
        #     print(f"Error: No sheets found in '{file_path}'.")
        #     return None

        if not first_sheet:
            # This case might be rare if workbook.active is used and the file is valid.
            print(f"Error: Could not access the first sheet in '{file_path}'.")
            return None

        print(f"Successfully accessed sheet: '{first_sheet.title}'")

        all_values = list(first_sheet.values)

        if not all_values:
            print(
                f"Sheet '{first_sheet.title}' in '{file_path}' is empty. Returning an empty DataFrame."
            )
            return pd.DataFrame()

        # Assume the first row is the header
        header = all_values[0]
        # The rest of the rows are data
        data_rows = all_values[1:]

        df = pd.DataFrame(data_rows, columns=header)
        print(
            f"Successfully read data from sheet '{first_sheet.title}' in '{file_path}' into a DataFrame using openpyxl."
        )
        return df

    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return None
    except Exception as e:
        # Handles other potential errors from openpyxl (e.g., corrupted file)
        print(f"An error occurred while processing '{file_path}': {e}")
        return None


def read_first_sheet_to_dataframe(file_path: str) -> Optional[pd.DataFrame]:
    """
    Reads all data from the first sheet of an Excel file into a pandas DataFrame.

    Args:
        file_path: The path to the .xlsx file.

    Returns:
        A pandas DataFrame containing the data from the first sheet.
        Returns None if the file is not found or an error occurs during reading.
    """
    try:
        # By default, read_excel reads the first sheet (sheet_name=0)
        # Specifying engine='openpyxl' is good practice for .xlsx files
        df = pd.read_excel(file_path, sheet_name=0, engine="openpyxl")
        print(f"Successfully read '{file_path}' into a DataFrame.")
        return df
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return None
    except ImportError:
        print(
            "Pandas or openpyxl library not found. Please install them: pip install pandas openpyxl"
        )
        return None
    except Exception as e:
        # Handles other potential errors from pandas/openpyxl (e.g., corrupted file, no sheets)
        print(f"An error occurred while reading '{file_path}' into a DataFrame: {e}")
        return None


# --- Example of how to use the function ---
if __name__ == "__main__":
    # Replace "your_excel_file.xlsx" with the actual path to your file.
    # For instance, you could use one of the files from your provided context:
    # example_file_path = "melbourne_2025_06_04_09_47DS.xlsx"

    # For demonstration, let's assume you have a file named "sample.xlsx"
    # in the same directory as your script.
    # You'll need to create this file or change the path.

    files_toparse = [
        "daily_overviews/melbourne_2025_06_04_09_47DS.xlsx",
        "daily_overviews/sydney_2025_06_04_09_48DS.xlsx",
        "daily_overviews/canberra_2025_06_04_09_48DS.xlsx",
        "daily_overviews/adelaide_2025_06_04_09_49DS.xlsx",
    ]

    # Create a dummy sample.xlsx for the example to run
    sheet_content = read_first_sheet_with_openpyxl(files_toparse[0])

    print(f"\nData from '{files_toparse[0]}':")
    print(sheet_content)
