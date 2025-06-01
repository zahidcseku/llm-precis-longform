"""
This script does the followings:
1. extacts daily forecasts form the JW API given location(s)
2. extracted the variables in  VARS from the API response
3. setup the promt for the LLM to generate a daily overview:
    - instructions
    - varible definitions
    - variables
4. send the prompt to the LLM
5. save the response to a file
"""

import dotenv
from var_dictionary import get_var_definitions
from utils import get_daily_forecasts, convert_to_tabular
from llm_utils import apply_llm
import json
from bom_scrapper import scrape_forecast_texts
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from datetime import datetime
from transformers import BertTokenizer, BertModel
from bert_score import BERTScorer

dotenv.load_dotenv()

"""
Setup:
1. LOCS={label: (lat, lon), ...}  # dictionary of tuples with latitude and longitude label will be used to save results
2. VARS=[var, ...]  # List of variable names to extract from the API response
"""
LOCS = {
    "melbourne": ("-37.8136", "144.9631"),
    "sydney": ("-33.86", "151.20"),
    "canberra": ("-35.31", "149.20"),
    "brisbane": ("-27.470125", "153.021072"),
    "perth": ("-31.95", "115.86"),
    "adelaide": ("-34.92", "138.59"),
    "hobart": ("-42.88", "147.33"),
    "darwin": ("-12.462827", "130.841782"),
}  # Example: Melbourne, SYD, CAN  Sydney Lat: -33.86Lon: 151.20 and Canberra Lat: -35.31Lon: 149.20
CITY_TO_STATE = {
    "melbourne": "vic",
    "sydney": "nsw",
    "canberra": "act",
    "brisbane": "qld",
    "perth": "wa",
    "adelaide": "sa",
    "hobart": "tas",
    "darwin": "nt",
}  #

VARS = [
    "cape_srf",
    "chill_stress_idx",
    "dew_pt",
    "frost_prob_cat",
    "tcc",
    "apparent_temp",
    "heat_stress_rating",
    "k_idx",
    "precip",
    "precip_conf",
    "precip_prob",
    "pres_msl",
    "rel_hum",
    "storm_prob_idx",
    "temp",
    "temp_inv_prob_idx",
    "total_totals_idx",
    "uv_idx_clear",
    "wind_dir",
    "wind_dir_compass",
    "wind_kmh",
]

VARS_SET = set(VARS)  # Use a set for efficient O(1) average time complexity lookups
# Prepare the variable definitions
var_definitions = get_var_definitions(VARS)


def bert_scorer(candidate, reference):
    """
    Calculate BERTScore for the candidate and reference texts.

    :param candidate: The generated text from the LLM.
    :param reference: The original text from BOM.
    :return: BERTScore F1 score.
    """
    scorer = BERTScorer(model_type="bert-base-uncased")
    P, R, F1 = scorer.score([candidate], [reference])

    return F1.mean().item()


def main(comments):
    for location_label, loc in LOCS.items():
        # get the bom daily forecasts for the location
        print(f"Processing daily forecasts for {location_label}...")
        # (state, city)
        bom_forecasts = scrape_forecast_texts(
            state=CITY_TO_STATE[location_label], city=location_label
        )

        # get the responses.
        data = get_daily_forecasts(loc, VARS, jw_model="access-g.13km")  # ai_enhanced

        """
        data is a dictionary with the following structure:
        "date": { "hour_minute": {var: value, ...}, ... }, ...}
        """
        if not data:
            print(f"No data found for {location_label}. Skipping...")
            continue

        # Convert the data to tabular format
        tabdata = convert_to_tabular(data)
        # Create a DataFrame from the tabular data
        data_df = pd.DataFrame(tabdata)

        # create a workbook for each location
        print(f"Processing location: {location_label} at coordinates {loc}...")
        workbook = Workbook()
        # Remove the default sheet created
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        sheet = workbook.create_sheet(title="overview")
        # Write the forecast text to the sheet
        sheet["A1"] = "Issued at:"
        sheet["B1"] = bom_forecasts["issued_at"]

        sheet["A3"] = "Date"
        sheet["B3"] = "bom_precis"
        sheet["C3"] = "llma_precis"
        sheet["D3"] = "bert_score"
        sheet["E3"] = "deepseek_precis"
        sheet["F3"] = "bert_score"
        sheet["G3"] = "mistral_precis"
        sheet["H3"] = "bert_score"
        sheet["I3"] = "gemini_precis"
        sheet["J3"] = "bert_score"

        sheet["L3"] = "bom_long_form"
        sheet["M3"] = "llma_long_form"
        sheet["N3"] = "bert_score"
        sheet["O3"] = "deepseek_long_form"
        sheet["P3"] = "bert_score"
        sheet["Q3"] = "mistral_long_form"
        sheet["R3"] = "bert_score"
        sheet["S3"] = "gemini_long_form"
        sheet["T3"] = "bert_score"

        # Add headers for the new columns
        sheet["V3"] = "llma_input_tokens"
        sheet["W3"] = "deepseek_input_tokens"
        sheet["X3"] = "mistral_input_tokens"
        sheet["Y3"] = "gemini_input_tokens"
        sheet["Z3"] = "llma_output_tokens"
        sheet["AA3"] = "deepseek_output_tokens"
        sheet["AB3"] = "mistral_output_tokens"
        sheet["AC3"] = "gemini_output_tokens"
        sheet["AD3"] = "llma_latency"
        sheet["AE3"] = "deepseek_latency"
        sheet["AF3"] = "mistral_latency"
        sheet["AG3"] = "gemini_latency"

        # token counts and latency

        # iterate for each date in the bom_forecasts
        row = 4
        for forecast_texts in bom_forecasts["daily_forecasts"]:
            current_day = f"{forecast_texts['day']} {datetime.now().year}"
            # %A for full weekday name, %d for day of the month, %B for full month name

            date_object = datetime.strptime(current_day, "%A %d %B %Y")

            # make the forecast data key
            forecast_key = date_object.strftime("%Y-%m-%d")
            # Check if the date exists in the data
            if forecast_key not in data:
                print(f"No data found for {forecast_key}. Skipping...")
                continue
            # filter the data for the current date
            hourly_data = data[forecast_key]
            # Apply the LLM to generate summaries
            llm_outputs = apply_llm(
                hourly_data,
                var_definitions,
            )

            # BERTScore calculation
            scorer = BERTScorer(model_type="bert-base-uncased")

            # Write the date and precis to the sheet
            sheet[f"A{row}"] = current_day

            sheet[f"B{row}"] = forecast_texts["precis"]
            sheet[f"C{row}"] = llm_outputs["llama-3.1-8b-instant"]["precis"]
            sheet[f"D{row}"] = bert_scorer(
                llm_outputs["llama-3.1-8b-instant"]["precis"],
                forecast_texts["precis"],
            )
            sheet[f"E{row}"] = llm_outputs["deepseek-r1-distill-llama-70b"]["precis"]
            sheet[f"F{row}"] = bert_scorer(
                llm_outputs["deepseek-r1-distill-llama-70b"]["precis"],
                forecast_texts["precis"],
            )
            sheet[f"G{row}"] = llm_outputs["mistral-saba-24b"]["precis"]
            sheet[f"H{row}"] = bert_scorer(
                llm_outputs["mistral-saba-24b"]["precis"],
                forecast_texts["precis"],
            )
            sheet[f"I{row}"] = llm_outputs["gemini-2.0-flash"]["precis"]
            sheet[f"J{row}"] = bert_scorer(
                llm_outputs["gemini-2.0-flash"]["precis"],
                forecast_texts["precis"],
            )

            sheet[f"L{row}"] = forecast_texts["long_form_text"]
            sheet[f"M{row}"] = llm_outputs["llama-3.1-8b-instant"]["long_form_text"]
            sheet[f"N{row}"] = bert_scorer(
                llm_outputs["llama-3.1-8b-instant"]["long_form_text"],
                forecast_texts["long_form_text"],
            )
            sheet[f"O{row}"] = llm_outputs["deepseek-r1-distill-llama-70b"][
                "long_form_text"
            ]
            sheet[f"P{row}"] = bert_scorer(
                llm_outputs["deepseek-r1-distill-llama-70b"]["long_form_text"],
                forecast_texts["long_form_text"],
            )
            sheet[f"Q{row}"] = llm_outputs["mistral-saba-24b"]["long_form_text"]
            sheet[f"R{row}"] = bert_scorer(
                llm_outputs["mistral-saba-24b"]["long_form_text"],
                forecast_texts["long_form_text"],
            )
            sheet[f"S{row}"] = llm_outputs["gemini-2.0-flash"]["long_form_text"]
            sheet[f"T{row}"] = bert_scorer(
                llm_outputs["gemini-2.0-flash"]["long_form_text"],
                forecast_texts["long_form_text"],
            )
            # Write the token counts and latency
            sheet[f"V{row}"] = llm_outputs["llama-3.1-8b-instant"]["input_tokens"]
            sheet[f"W{row}"] = llm_outputs["deepseek-r1-distill-llama-70b"][
                "input_tokens"
            ]
            sheet[f"X{row}"] = llm_outputs["mistral-saba-24b"]["input_tokens"]
            sheet[f"Y{row}"] = llm_outputs["gemini-2.0-flash"]["input_tokens"]
            sheet[f"Z{row}"] = llm_outputs["llama-3.1-8b-instant"]["output_tokens"]
            sheet[f"AA{row}"] = llm_outputs["deepseek-r1-distill-llama-70b"][
                "output_tokens"
            ]
            sheet[f"AB{row}"] = llm_outputs["mistral-saba-24b"]["output_tokens"]
            sheet[f"AC{row}"] = llm_outputs["gemini-2.0-flash"]["output_tokens"]
            sheet[f"AD{row}"] = llm_outputs["llama-3.1-8b-instant"]["latency_seconds"]
            sheet[f"AE{row}"] = llm_outputs["deepseek-r1-distill-llama-70b"][
                "latency_seconds"
            ]
            sheet[f"AF{row}"] = llm_outputs["mistral-saba-24b"]["latency_seconds"]
            sheet[f"AG{row}"] = llm_outputs["gemini-2.0-flash"]["latency_seconds"]

            # Write the hourly data to a new sheet
            day_sheet = workbook.create_sheet(title=forecast_key)
            day_df = data_df[data_df["date"] == forecast_key]

            # Write headers to the new sheet
            headers = list(day_df.columns)
            day_sheet.append(headers)

            # Write data rows to the new sheet
            for r_idx, row_data in enumerate(
                day_df.values.tolist(), 1
            ):  # openpyxl is 1-indexed for rows if not using append
                # day_sheet.append(row_data) # Alternative using append
                for c_idx, value in enumerate(row_data, 1):
                    try:
                        day_sheet.cell(
                            row=r_idx + 1, column=c_idx, value=value
                        )  # +1 for r_idx because headers are on row 1
                    except IllegalCharacterError:
                        day_sheet.cell(
                            row=r_idx + 1, column=c_idx, value="[ILLEGAL CHAR]"
                        )  # Placeholder for illegal chars

            row += 1

        # write comments
        sheet[f"A{row + 3}"] = comments

        # Save the workbook
        output_filename = f"daily_overviews/{location_label}_{datetime.now().strftime('%Y_%m_%d_%H_%M')}.xlsx"
        workbook.save(output_filename)


if __name__ == "__main__":
    comments = "prompt including Jane's comments no BoM data yet"
    main(comments)
